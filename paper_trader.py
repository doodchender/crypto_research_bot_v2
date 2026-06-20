import asyncio
import json
import logging
import math
import sqlite3
import time
from datetime import datetime
from pathlib import Path
from typing import Optional, Callable

from risk_manager import RiskManager

logger = logging.getLogger("paper")

def init_paper_db(db_path: str):
    conn = sqlite3.connect(db_path)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS paper_trades (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            timestamp_open  TEXT NOT NULL,
            timestamp_close TEXT,
            symbol          TEXT NOT NULL,   -- btc / eth
            side            TEXT NOT NULL,   -- long / short
            entry_price     REAL NOT NULL,
            exit_price      REAL,
            size_usd        REAL NOT NULL,
            size_pct        REAL NOT NULL,
            sl_pct          REAL NOT NULL,
            tp_pct          REAL NOT NULL,
            predicted_pct   REAL,            -- прогноз модели в %
            actual_pct      REAL,            -- фактическое движение в %
            pnl_usd         REAL,
            pnl_pct         REAL,
            fees_usd        REAL,
            reason          TEXT,            -- TP / SL / TIMEOUT / MANUAL
            status          TEXT NOT NULL,   -- open / closed
            features_json   TEXT,            -- снимок ключевых фич
            capital_before  REAL,
            capital_after   REAL
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_paper_status ON paper_trades(status)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_paper_ts ON paper_trades(timestamp_open)")

    conn.execute("""
        CREATE TABLE IF NOT EXISTS paper_state (
            key TEXT PRIMARY KEY,
            value REAL
        )
    """)
    conn.commit()
    conn.close()

def _load_capital(db_path: str, default: float) -> float:
    try:
        conn = sqlite3.connect(db_path)
        row = conn.execute("SELECT value FROM paper_state WHERE key='capital'").fetchone()
        conn.close()
        if row and row[0] is not None:
            return float(row[0])
    except Exception:
        pass
    return default

def _save_capital(db_path: str, cap: float):
    try:
        conn = sqlite3.connect(db_path)
        conn.execute("INSERT OR REPLACE INTO paper_state(key, value) VALUES ('capital', ?)", (cap,))
        conn.commit()
        conn.close()
    except Exception as e:
        logger.error(f"save_capital: {e}")

class PaperTrader:

    INTERVAL_SEC = 300
    FEE_PCT = 0.04
    MIN_PREDICTED_PCT = 0.10
    POSITION_TIMEOUT_CANDLES = 6

    def __init__(
        self,
        intraday_db: str,
        paper_db: str,
        initial_capital: float = 1000.0,
        notifier: Optional[Callable] = None,
        logger: Optional[logging.Logger] = None,
    ):
        self.intraday_db = intraday_db
        self.paper_db = paper_db
        self.initial_capital = initial_capital
        self.notifier = notifier
        self.logger = logger or globals()["logger"]

        init_paper_db(paper_db)
        loaded = _load_capital(paper_db, initial_capital)
        self.risk = RiskManager(initial_capital=loaded)
        self.risk.current_capital = loaded
        self.running = False
        self.current_day = None

    async def _notify(self, text: str):
        if self.notifier is None:
            return
        try:
            await self.notifier(text)
        except Exception as e:
            self.logger.error(f"paper notify: {e}")

    def _open_position_record(self, data: dict) -> int:
        conn = sqlite3.connect(self.paper_db)
        cur = conn.execute("""
            INSERT INTO paper_trades (
                timestamp_open, symbol, side, entry_price, size_usd, size_pct,
                sl_pct, tp_pct, predicted_pct, status, features_json, capital_before
            ) VALUES (?,?,?,?,?,?,?,?,?,'open',?,?)
        """, (
            data["ts"], data["symbol"], data["side"], data["entry_price"],
            data["size_usd"], data["size_pct"], data["sl_pct"], data["tp_pct"],
            data["predicted_pct"], json.dumps(data["features"], default=float),
            data["capital"],
        ))
        trade_id = cur.lastrowid
        conn.commit()
        conn.close()
        return trade_id

    def _close_position_record(self, trade_id: int, exit_price: float,
                                pnl_usd: float, pnl_pct: float, reason: str,
                                fees: float, actual_pct: float, ts_close: str,
                                capital_after: float):
        conn = sqlite3.connect(self.paper_db)
        conn.execute("""
            UPDATE paper_trades
            SET timestamp_close=?, exit_price=?, pnl_usd=?, pnl_pct=?,
                fees_usd=?, reason=?, actual_pct=?, status='closed', capital_after=?
            WHERE id=?
        """, (ts_close, exit_price, pnl_usd, pnl_pct, fees, reason, actual_pct,
              capital_after, trade_id))
        conn.commit()
        conn.close()

    def _get_open_positions(self) -> list[dict]:
        conn = sqlite3.connect(self.paper_db)
        conn.row_factory = sqlite3.Row
        rows = conn.execute("SELECT * FROM paper_trades WHERE status='open'").fetchall()
        conn.close()
        return [dict(r) for r in rows]

    async def _tick(self):
        from intraday_forecast import get_intraday_forecast

        ts_now = datetime.now().strftime("%Y-%m-%d %H:%M")
        day = ts_now[:10]
        if day != self.current_day:
            self.current_day = day
            self.risk.start_new_day(day)

        try:
            conn = sqlite3.connect(self.intraday_db)
            latest = conn.execute("""
                SELECT timestamp, btc_price, eth_price
                FROM intraday_snapshots
                WHERE btc_price IS NOT NULL
                ORDER BY timestamp DESC LIMIT 1
            """).fetchone()
            conn.close()
        except Exception as e:
            self.logger.debug(f"paper: cannot read intraday {e}")
            return

        if not latest:
            return
        snap_ts, btc_price, eth_price = latest

        for pos in self._get_open_positions():
            sym = pos["symbol"]
            cur_price = btc_price if sym == "btc" else eth_price
            if not cur_price:
                continue

            entry = pos["entry_price"]
            side = pos["side"]
            sign = 1 if side == "long" else -1
            pnl_pct_move = (cur_price - entry) / entry * sign

            try:
                t_open = datetime.strptime(pos["timestamp_open"], "%Y-%m-%d %H:%M")
                t_now = datetime.strptime(snap_ts, "%Y-%m-%d %H:%M")
                candles_held = max(1, int((t_now - t_open).total_seconds() / 300))
            except Exception:
                candles_held = 1

            hit_tp = pnl_pct_move >= pos["tp_pct"]
            hit_sl = pnl_pct_move <= -pos["sl_pct"]
            timeout = candles_held >= self.POSITION_TIMEOUT_CANDLES

            if hit_tp or hit_sl or timeout:
                reason = "TP" if hit_tp else ("SL" if hit_sl else "TIMEOUT")
                size_usd = pos["size_usd"]
                gross = size_usd * pnl_pct_move
                fees = size_usd * (self.FEE_PCT / 100) * 2
                net = gross - fees
                self.risk.register_trade_pnl(net)
                _save_capital(self.paper_db, self.risk.current_capital)

                actual_pct = pnl_pct_move * 100
                self._close_position_record(
                    pos["id"], cur_price, net, actual_pct, reason, fees,
                    actual_pct, snap_ts, self.risk.current_capital
                )

                emoji = "✅" if net > 0 else "❌"
                pred = pos["predicted_pct"] or 0
                pred_match = "✓" if (pred > 0) == (net > 0) else "✗"

                status = self.risk.get_status()
                text = (
                    f"{emoji} <b>PAPER TRADE CLOSED</b> [{reason}]\n"
                    f"<code>{sym.upper()} {side.upper()}</code>\n"
                    f"${entry:,.2f} → ${cur_price:,.2f} ({actual_pct:+.3f}%)\n"
                    f"Удержано: {candles_held * 5} мин\n"
                    f"P&L: <code>${net:+.2f}</code> (fees ${fees:.2f})\n"
                    f"Прогноз/Факт: {pred:+.3f}% / {actual_pct:+.3f}% {pred_match}\n\n"
                    f"💰 Капитал: <code>${self.risk.current_capital:,.2f}</code> "
                    f"({status.get('total_return_pct', 0):+.2f}%)\n"
                    f"📊 За сегодня: {status.get('trades_today', 0)} сделок, "
                    f"WR {status.get('win_rate_today', 0):.0f}%"
                )
                await self._notify(text)
                self.logger.info(f"paper closed {sym} {side} {reason} pnl=${net:+.2f}")

        open_symbols = {p["symbol"] for p in self._get_open_positions()}

        for sym in ["btc", "eth"]:
            if sym in open_symbols:
                continue

            try:
                forecast = get_intraday_forecast(sym, self.intraday_db)
            except Exception as e:
                self.logger.debug(f"paper forecast {sym}: {e}")
                continue

            if "error" in forecast:
                continue

            pred_pct = float(forecast.get("predicted_pct", 0))
            if abs(pred_pct) < self.MIN_PREDICTED_PCT:
                continue

            side = "long" if pred_pct > 0 else "short"
            sign = 1 if side == "long" else -1
            features = forecast.get("inputs", {}) or {}

            ob_imb = float(features.get("ob_imb") or 0)
            ret_lag1 = float(features.get("ret_lag1") or 0)
            flow_br = float(features.get("flow_buy_ratio") or 0)

            ob_ok = (ob_imb * sign) > 0.1
            mom_ok = (ret_lag1 * sign) > 0
            flow_ok = (flow_br * sign) >= 0

            if not (ob_ok and mom_ok):
                continue

            decision = self.risk.evaluate_entry(
                predicted_pct=pred_pct,
                side=side,
                features=features,
            )
            if not decision["allow"]:
                self.logger.debug(f"paper {sym} declined: {decision['reason']}")
                continue

            cur_price = btc_price if sym == "btc" else eth_price
            trade_id = self._open_position_record({
                "ts": snap_ts,
                "symbol": sym,
                "side": side,
                "entry_price": cur_price,
                "size_usd": decision["size_usd"],
                "size_pct": decision["size_pct"],
                "sl_pct": decision["sl_pct"],
                "tp_pct": decision["tp_pct"],
                "predicted_pct": pred_pct,
                "capital": self.risk.current_capital,
                "features": {
                    "ob_imb": ob_imb,
                    "ret_lag1": ret_lag1,
                    "whale_imb": features.get("whale_imb") or 0,
                    "social_sentiment": features.get("social_sentiment") or 0,
                    "cascade_risk": features.get("cascade_risk") or 0,
                    "realized_vol": features.get("realized_vol") or 0,
                    "egarch_vol": features.get("egarch_vol") or 0,
                    "fear_greed": features.get("fear_greed") or 0,
                },
            })

            arrow = "📈" if side == "long" else "📉"
            cascade = features.get("cascade_risk") or 0
            warn = ""
            if cascade > 0.4:
                warn = f"\n⚠️ <i>Cascade risk: {cascade:.2f}</i>"

            text = (
                f"{arrow} <b>PAPER TRADE OPENED</b> #{trade_id}\n"
                f"<code>{sym.upper()} {side.upper()}</code>\n"
                f"Entry: <code>${cur_price:,.2f}</code>\n"
                f"Размер: <code>${decision['size_usd']:.2f}</code> "
                f"({decision['size_pct']*100:.1f}% от капитала)\n"
                f"SL: <code>{-decision['sl_pct']*100:.2f}%</code>  "
                f"TP: <code>+{decision['tp_pct']*100:.2f}%</code>\n"
                f"Прогноз: <code>{pred_pct:+.3f}%</code>\n"
                f"Boost: ×{decision['boost_multiplier']:.2f}"
                f"{warn}"
            )
            await self._notify(text)
            self.logger.info(
                f"paper opened {sym} {side} @ {cur_price} size=${decision['size_usd']:.2f}"
            )

    async def run(self):
        self.running = True
        self.logger.info("Paper trader started")

        await asyncio.sleep(30)

        while self.running:
            try:
                await self._tick()
            except Exception as e:
                self.logger.error(f"paper tick error: {e}")
            await asyncio.sleep(self.INTERVAL_SEC)

    def stop(self):
        self.running = False

    def reset_capital(self, new_capital: float = None):
        cap = new_capital if new_capital is not None else self.initial_capital
        self.risk.current_capital = cap
        self.risk.peak_capital = cap
        self.risk.daily = None
        _save_capital(self.paper_db, cap)

    def generate_excel_report(self, out_path: str) -> str:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        conn = sqlite3.connect(self.paper_db)
        conn.row_factory = sqlite3.Row
        trades = [dict(r) for r in conn.execute(
            "SELECT * FROM paper_trades ORDER BY timestamp_open"
        ).fetchall()]
        conn.close()

        wb = Workbook()
        ws = wb.active
        ws.title = "Trades"

        headers = [
            "ID", "Открыто", "Закрыто", "Актив", "Сторона",
            "Цена входа", "Цена выхода", "Размер $", "Размер %",
            "Прогноз %", "Факт %", "Совпал?", "P&L $", "P&L %",
            "Fees $", "SL %", "TP %", "Причина",
            "Cascade", "Vol", "Whale imb", "Sentiment", "Капитал после"
        ]
        ws.append(headers)

        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        green_fill = PatternFill("solid", fgColor="C6EFCE")
        red_fill = PatternFill("solid", fgColor="FFC7CE")
        yellow_fill = PatternFill("solid", fgColor="FFEB9C")

        for t in trades:
            try:
                feat = json.loads(t.get("features_json") or "{}")
            except Exception:
                feat = {}

            pred = t.get("predicted_pct") or 0
            actual = t.get("actual_pct") or 0
            pnl = t.get("pnl_usd") or 0
            direction_match = "✓" if (pred > 0) == (actual > 0) and t.get("status") == "closed" else ("" if t.get("status") == "open" else "✗")

            row = [
                t["id"],
                t["timestamp_open"],
                t.get("timestamp_close", ""),
                (t["symbol"] or "").upper(),
                (t["side"] or "").upper(),
                t["entry_price"],
                t.get("exit_price", ""),
                round(t["size_usd"], 2),
                round((t["size_pct"] or 0) * 100, 2),
                round(pred, 3),
                round(actual, 3) if t.get("status") == "closed" else "",
                direction_match,
                round(pnl, 2) if t.get("status") == "closed" else "",
                round(t.get("pnl_pct") or 0, 3) if t.get("status") == "closed" else "",
                round(t.get("fees_usd") or 0, 2) if t.get("status") == "closed" else "",
                round(t["sl_pct"] * 100, 3),
                round(t["tp_pct"] * 100, 3),
                t.get("reason") or "open",
                round(float(feat.get("cascade_risk", 0)), 3),
                round(float(feat.get("egarch_vol", feat.get("realized_vol", 0))), 5),
                round(float(feat.get("whale_imb", 0)), 3),
                round(float(feat.get("social_sentiment", 0)), 3),
                round(t.get("capital_after") or 0, 2) if t.get("status") == "closed" else "",
            ]
            ws.append(row)

            row_idx = ws.max_row
            if t.get("status") == "closed":
                fill = green_fill if pnl > 0 else red_fill
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col).fill = fill
            else:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col).fill = yellow_fill

        widths = [5, 17, 17, 6, 7, 11, 11, 10, 10, 10, 10, 9, 10, 10, 9, 8, 8, 10, 9, 9, 10, 11, 13]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.row_dimensions[1].height = 30

        ws.freeze_panes = "A2"

        ws2 = wb.create_sheet("Summary")
        closed_trades = [t for t in trades if t.get("status") == "closed"]
        wins = [t for t in closed_trades if (t.get("pnl_usd") or 0) > 0]
        losses = [t for t in closed_trades if (t.get("pnl_usd") or 0) <= 0]

        total_pnl = sum((t.get("pnl_usd") or 0) for t in closed_trades)
        total_fees = sum((t.get("fees_usd") or 0) for t in closed_trades)
        win_rate = len(wins) / len(closed_trades) * 100 if closed_trades else 0
        avg_win = sum((t.get("pnl_usd") or 0) for t in wins) / len(wins) if wins else 0
        avg_loss = sum((t.get("pnl_usd") or 0) for t in losses) / len(losses) if losses else 0

        direction_matches = sum(
            1 for t in closed_trades
            if ((t.get("predicted_pct") or 0) > 0) == ((t.get("actual_pct") or 0) > 0)
        )
        direction_pct = direction_matches / len(closed_trades) * 100 if closed_trades else 0

        summary = [
            ("Начальный капитал", self.initial_capital),
            ("Текущий капитал", round(self.risk.current_capital, 2)),
            ("Итого P&L", round(total_pnl, 2)),
            ("Доходность %", round((self.risk.current_capital - self.initial_capital) /
                                    self.initial_capital * 100, 2)),
            ("Комиссии", round(total_fees, 2)),
            ("", ""),
            ("Всего сделок (закрытых)", len(closed_trades)),
            ("Открытых сейчас", len(trades) - len(closed_trades)),
            ("Прибыльных", len(wins)),
            ("Убыточных", len(losses)),
            ("Win rate %", round(win_rate, 1)),
            ("Точность направления %", round(direction_pct, 1)),
            ("Ср. выигрыш $", round(avg_win, 2)),
            ("Ср. убыток $", round(avg_loss, 2)),
            ("R/R ratio", round(abs(avg_win / avg_loss), 2) if avg_loss else 0),
        ]
        ws2.append(["Метрика", "Значение"])
        for m, v in summary:
            ws2.append([m, v])
        ws2.column_dimensions["A"].width = 30
        ws2.column_dimensions["B"].width = 15
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font

        wb.save(out_path)
        return out_path

    def get_stats_text(self) -> str:
        conn = sqlite3.connect(self.paper_db)
        conn.row_factory = sqlite3.Row
        closed = [dict(r) for r in conn.execute(
            "SELECT * FROM paper_trades WHERE status='closed'"
        ).fetchall()]
        open_trades = [dict(r) for r in conn.execute(
            "SELECT * FROM paper_trades WHERE status='open'"
        ).fetchall()]
        conn.close()

        n = len(closed)
        wins = sum(1 for t in closed if (t.get("pnl_usd") or 0) > 0)
        total = sum((t.get("pnl_usd") or 0) for t in closed)
        wr = wins / n * 100 if n > 0 else 0

        status = self.risk.get_status()

        lines = [
            "📊 <b>Paper Trading — статистика</b>",
            "",
            f"💰 Капитал: <code>${status['capital']:,.2f}</code> "
            f"({status['total_return_pct']:+.2f}%)",
            f"🏔 Peak: <code>${status['peak']:,.2f}</code>",
            "",
            f"📈 Всего сделок: <b>{n}</b> (открытых: {len(open_trades)})",
            f"✅ Win rate: <b>{wr:.0f}%</b> ({wins}/{n})",
            f"💵 Итого P&L: <code>${total:+.2f}</code>",
        ]
        if self.risk.daily:
            lines += [
                "",
                f"📅 За сегодня ({self.risk.daily.date}):",
                f"  Сделок: {self.risk.daily.trades_count}",
                f"  P&L: <code>${self.risk.daily.realized_pnl:+.2f}</code>",
                f"  DD: <code>{self.risk.daily.daily_drawdown*100:+.2f}%</code>",
                f"  Уровень: {self.risk.daily.level_reached}/3",
            ]
        if open_trades:
            lines += ["", "🔓 <b>Открытые позиции:</b>"]
            for p in open_trades:
                lines.append(
                    f"  #{p['id']}: {(p['symbol'] or '').upper()} "
                    f"{(p['side'] or '').upper()} @ ${p['entry_price']:,.2f} "
                    f"(${p['size_usd']:.2f})"
                )
        return "\n".join(lines)
